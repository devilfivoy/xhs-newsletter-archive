#!/usr/bin/env python3
"""
product-showcase-page: 从Excel货盘表生成产品展示HTML页面

用法:
    python generate_showcase.py \
        --excel 货盘表.xlsx \
        --title "卤味果干类新品推荐" \
        --output ./output \
        --categories '{"卤味": ["鸭脖","鸭舌","猪肝片"], "坚果果干零食": ["桂圆","西梅"]}' \
        --product-images '{"商品名": "图片URL"}' \
        --hancards '{"商品名": "手卡URL"}'

依赖: openpyxl, Pillow, requests
"""

import argparse
import json
import os
import re
import html as h
from typing import Dict, List, Optional, Tuple


def parse_excel(path: str, skip_rows: Optional[List[int]] = None) -> List[dict]:
    """解析Excel货盘表，返回商品列表"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    
    products = []
    for row in range(2, ws.max_row + 1):
        if skip_rows and row in skip_rows:
            continue
        name = ws.cell(row, 4).value
        if not name:
            continue
        products.append({
            'am': ws.cell(row, 1).value or '',
            'shop_id': ws.cell(row, 2).value or '',
            'shop': ws.cell(row, 3).value or '',
            'name': str(name).strip(),
            'item_id': ws.cell(row, 5).value or '',
            'link': ws.cell(row, 6).value or '',
            'diff': ws.cell(row, 8).value or '',
            'row': row,
        })
    return products


def extract_keywords(diff_text: str, custom_patterns: Optional[List[str]] = None) -> List[str]:
    """从差异点文本中提取3个核心关键词"""
    if not diff_text:
        return ['源头直供', '品质保障', '新品推荐']
    
    text = str(diff_text)
    default_patterns = [
        r'高蛋白', r'0脂肪', r'0蔗糖', r'非油炸', r'独立包装',
        r'低温慢烘', r'无添加', r'无防腐剂', r'原切', r'手工',
        r'富硒', r'药食同源', r'膳食纤维', r'配料干净',
        r'先卤后烤', r'先卤后熏', r'铁棍山药', r'鲜蒸',
        r'九成风干', r'即食', r'短保', r'老卤慢煮',
        r'果木柴火', r'控温烘烤', r'酥脆', r'鲜果原切',
        r'老醋', r'传统酿造', r'解腻开胃', r'椰浆',
        r'无蔗糖', r'天然代糖', r'0色素', r'0添加剂',
    ]
    
    patterns = (custom_patterns or []) + default_patterns
    keywords = []
    seen = set()
    
    for pat in patterns:
        if len(keywords) >= 3:
            break
        m = re.search(pat, text)
        if m and m.group() not in seen:
            keywords.append(m.group())
            seen.add(m.group())
    
    while len(keywords) < 3:
        keywords.append('新品推荐')
    return keywords[:3]


def clean_diff(text: str) -> str:
    """清洗差异点文本为统一有序编号格式"""
    if not text:
        return ''
    lines = str(text).split('\n')
    items = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        line = re.sub(r'^[①②③④⑤⑥⑦⑧⑨⑩✅▪️√·•\-—]+\s*', '', line)
        line = re.sub(r'^[\d]+[、\.．\)）]\s*', '', line)
        line = re.sub(r'^[一二三四五六七八九十]+[、\.．：:]\s*', '', line)
        line = re.sub(r'^\【[^】]*\】\s*', '', line)
        line = line.strip()
        if line:
            items.append(line)
    
    merged = []
    for item in items:
        if merged and len(item) < 10:
            merged[-1] += '，' + item
        else:
            merged.append(item)
    
    return '\n'.join(f"{i+1}. {item}" for i, item in enumerate(merged))


def compress_image(input_path: str, output_path: str, 
                   max_width: int = 800, quality: int = 55) -> int:
    """压缩图片，返回压缩后文件大小"""
    from PIL import Image
    img = Image.open(input_path)
    if img.mode == 'RGBA':
        img = img.convert('RGB')
    if img.width > max_width:
        ratio = max_width / img.width
        img = img.resize((max_width, int(img.height * ratio)), Image.LANCZOS)
    img.save(output_path, format='JPEG', quality=quality, optimize=True)
    return os.path.getsize(output_path)


def generate_card_html(product: dict, truncate_len: int = 100) -> str:
    """生成单个商品卡片HTML"""
    esc_name = h.escape(product['name'])
    esc_shop = h.escape(product['shop'])
    esc_link = h.escape(product['link'])
    cat = product.get('category', '未分类')
    
    tags = ''.join(f'<span class="tag">{h.escape(kw)}</span>' for kw in product['keywords'])
    
    diff_full = h.escape(product['diff_clean']).replace('\n', '<br>')
    if len(product['diff_clean']) > truncate_len:
        diff_short = h.escape(product['diff_clean'][:truncate_len]).replace('\n', '<br>') + '...'
        diff_html = f'''<div class="diff-text">
          <div class="short">{diff_short}</div>
          <div class="full" style="display:none">{diff_full}</div>
          <span class="expand-btn" onclick="var s=this.parentElement;s.querySelector('.short').style.display=s.querySelector('.short').style.display==='none'?'block':'none';s.querySelector('.full').style.display=s.querySelector('.full').style.display==='none'?'block':'none';this.textContent=this.textContent==='▼ 展开'?'▲ 收起':'▼ 展开';">▼ 展开</span>
        </div>'''
    else:
        diff_html = f'<div class="diff-text">{diff_full}</div>'
    
    hancard_html = ''
    if product.get('hancard_url'):
        hancard_html = f'''<div class="hancard-section">
          <div class="hancard-label">📋 产品手卡</div>
          <img src="{h.escape(product['hancard_url'])}" alt="手卡" class="hancard-img" loading="lazy" onclick="window.open(this.src)">
        </div>'''
    
    return f'''
    <div class="card" data-cat="{cat}">
      <div class="img-wrapper">
        <img src="{h.escape(product.get('img', ''))}" alt="{esc_name}" loading="lazy" onclick="window.open(this.src)">
        <span class="shop-badge">{esc_shop}</span>
      </div>
      <div class="card-body">
        <h3>{esc_name}</h3>
        <div class="tags">{tags}</div>
        {diff_html}
        {hancard_html}
        <a href="{esc_link}" target="_blank" class="btn">查看商品 →</a>
      </div>
    </div>'''


def generate_full_html(products: List[dict], title: str, 
                       categories: Optional[Dict[str, int]] = None) -> str:
    """生成完整HTML页面"""
    cards_html = ''.join(generate_card_html(p) for p in products)
    
    # 构建筛选按钮
    if categories:
        filter_btns = f'<button class="filter-btn active" onclick="filterCards(\'all\', this)">全部 ({len(products)})</button>'
        for cat, count in categories.items():
            filter_btns += f'<button class="filter-btn" onclick="filterCards(\'{cat}\', this)">{cat} ({count})</button>'
    else:
        filter_btns = ''
    
    filter_bar = f'<div class="filter-bar">{filter_btns}</div>' if filter_btns else ''
    
    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{h.escape(title)}</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#f5f7f5; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Hiragino Sans GB",sans-serif; color:#333; }}
.header {{ background:linear-gradient(135deg,#a8d5a2 0%,#7bc67e 50%,#5ab55e 100%); padding:28px 20px 20px; text-align:center; color:#fff; width:100%; }}
.header h1 {{ font-size:26px; font-weight:700; letter-spacing:4px; }}
.filter-bar {{ display:flex; justify-content:center; gap:10px; padding:14px 20px; background:#fff; border-bottom:1px solid #e8ede8; position:sticky; top:0; z-index:10; }}
.filter-btn {{ padding:6px 20px; border-radius:20px; border:1.5px solid #c8dcc8; background:#fff; color:#555; font-size:13px; font-weight:500; cursor:pointer; transition:all 0.2s; }}
.filter-btn:hover {{ border-color:#7bc67e; color:#4a9e4e; }}
.filter-btn.active {{ background:linear-gradient(135deg,#5ab55e,#4a9e4e); color:#fff; border-color:transparent; }}
.grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(220px,1fr)); gap:12px; padding:14px; max-width:1200px; margin:0 auto; }}
.card {{ background:#fff; border-radius:10px; overflow:hidden; box-shadow:0 1px 6px rgba(0,0,0,0.06); transition:transform 0.2s,box-shadow 0.2s; }}
.card:hover {{ transform:translateY(-3px); box-shadow:0 4px 16px rgba(74,158,78,0.12); }}
.card.hidden {{ display:none; }}
.img-wrapper {{ position:relative; width:100%; padding-top:100%; background:#f9f9f9; overflow:hidden; cursor:pointer; }}
.img-wrapper img {{ position:absolute; top:0; left:0; width:100%; height:100%; object-fit:cover; }}
.shop-badge {{ position:absolute; bottom:8px; left:8px; background:rgba(0,0,0,0.55); color:#fff; font-size:9px; padding:2px 6px; border-radius:3px; backdrop-filter:blur(4px); }}
.card-body {{ padding:10px 12px 12px; }}
.card-body h3 {{ font-size:13px; font-weight:600; margin-bottom:6px; line-height:1.4; color:#222; }}
.tags {{ display:flex; flex-wrap:wrap; gap:4px; margin-bottom:8px; }}
.tag {{ background:#e8f5e9; color:#2e7d32; font-size:9px; padding:2px 6px; border-radius:3px; font-weight:500; }}
.diff-text {{ font-size:10px; color:#666; line-height:1.5; margin-bottom:8px; }}
.expand-btn {{ color:#4a9e4e; font-size:10px; cursor:pointer; display:inline-block; margin-top:2px; }}
.expand-btn:hover {{ text-decoration:underline; }}
.hancard-section {{ margin-bottom:8px; }}
.hancard-label {{ font-size:9px; color:#8B6914; background:#FFF8DC; padding:2px 6px; border-radius:3px; display:inline-block; margin-bottom:4px; }}
.hancard-img {{ width:100%; border-radius:6px; cursor:pointer; border:1px solid #eee; }}
.btn {{ display:inline-block; background:linear-gradient(135deg,#5ab55e,#4a9e4e); color:#fff; text-decoration:none; font-size:10px; padding:5px 12px; border-radius:5px; font-weight:500; transition:opacity 0.2s; }}
.btn:hover {{ opacity:0.85; }}
.footer {{ text-align:center; padding:20px; font-size:10px; color:#bbb; }}
@media (max-width:600px) {{ .grid {{ grid-template-columns:repeat(2,1fr); gap:8px; padding:8px; }} .header h1 {{ font-size:20px; }} .filter-btn {{ font-size:11px; padding:5px 14px; }} }}
</style>
</head>
<body>
<div class="header"><h1>{h.escape(title)}</h1></div>
{filter_bar}
<div class="grid">{cards_html}</div>
<div class="footer">{h.escape(title)}</div>
<script>
function filterCards(cat, btn) {{
  document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  document.querySelectorAll('.card').forEach(card => {{
    card.classList.toggle('hidden', cat !== 'all' && card.dataset.cat !== cat);
  }});
}}
</script>
</body>
</html>'''


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='从Excel生成产品展示HTML')
    parser.add_argument('--excel', required=True, help='Excel货盘表路径')
    parser.add_argument('--title', default='新品推荐', help='页面标题')
    parser.add_argument('--output', default='./output', help='输出目录')
    parser.add_argument('--categories', help='分类映射JSON')
    parser.add_argument('--product-images', help='商品主图映射JSON')
    parser.add_argument('--hancards', help='手卡映射JSON')
    args = parser.parse_args()
    
    products = parse_excel(args.excel)
    cat_map = json.loads(args.categories) if args.categories else {}
    img_map = json.loads(args.product_images) if args.product_images else {}
    hc_map = json.loads(args.hancards) if args.hancards else {}
    
    for p in products:
        p['keywords'] = extract_keywords(p['diff'])
        p['diff_clean'] = clean_diff(p['diff'])
        p['img'] = img_map.get(p['name'], '')
        p['hancard_url'] = hc_map.get(p['name'])
        p['category'] = '未分类'
        for cat, names in cat_map.items():
            if p['name'] in names:
                p['category'] = cat
                break
    
    categories = {}
    for p in products:
        cat = p['category']
        categories[cat] = categories.get(cat, 0) + 1
    
    os.makedirs(args.output, exist_ok=True)
    html = generate_full_html(products, args.title, categories)
    
    output_path = os.path.join(args.output, 'index.html')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"✅ 已生成: {output_path} ({os.path.getsize(output_path)/1024:.1f}KB)")
    print(f"   商品数: {len(products)}")
    print(f"   分类: {categories}")
